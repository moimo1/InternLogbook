import os
import json
import time
import hashlib
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from flask import Flask, render_template, request, redirect, url_for, session, flash, g
from core.database import get_db_connection, init_db
from core.engine import calculate_intern_hours, get_schedule_for_date

def get_local_now():
    """
    Returns a naive datetime object representing the current date and time
    in the configured timezone (falls back to Asia/Manila).
    Using naive datetimes prevents comparison issues with database logs.
    """
    tz_name = os.environ.get('TZ', os.environ.get('APP_TZ', 'Asia/Manila'))
    try:
        return datetime.now(ZoneInfo(tz_name)).replace(tzinfo=None)
    except Exception:
        return datetime.now()

def get_current_qr_tokens():
    """
    Generates time-based tokens for the current and previous minute.
    This gives the user a 60-120 second window to scan and submit.
    """
    secret = app.secret_key or b'default_fallback_key'
    if isinstance(secret, str):
        secret = secret.encode('utf-8')
    current_minute = int(time.time() // 60)
    
    token_current = hashlib.sha256(secret + f":{current_minute}".encode('utf-8')).hexdigest()[:10]
    token_previous = hashlib.sha256(secret + f":{current_minute - 1}".encode('utf-8')).hexdigest()[:10]
    
    return [token_current, token_previous]

app = Flask(__name__)
# Secure fallback for missing production keys to prevent spoofing
app.secret_key = os.environ.get('SECRET_KEY') or os.urandom(24)

@app.teardown_appcontext
def close_db_connection(exception):
    db_conn = g.pop('db_conn', None)
    if db_conn is not None:
        try:
            db_conn.close()
        except Exception:
            pass

try:
    init_db()
except Exception as e:
    print(f"Skipping database schema initialization: {e}")

# -------------------------------------------------------------------------
# GLOBAL SESSION SECURITY INTERCEPTOR
# -------------------------------------------------------------------------
@app.before_request
def check_user_activity_status():
    """
    Executes globally before any endpoint router. Instantly revokes session 
    tokens and clears browser cookies if an administrative action deactivates 
    the current logged-in intern profile.
    """
    exempt_routes = ['log_in', 'register', 'static']
    if request.endpoint in exempt_routes:
        return None

    if 'user_id' in session:
        uid = session['user_id']
        
        with get_db_connection() as conn:
            from psycopg2.extras import RealDictCursor
            with conn.cursor(cursor_factory=RealDictCursor) as cursor:
                cursor.execute('SELECT is_active, roles FROM users WHERE id = %s', (uid,))
                user = cursor.fetchone()

        if not user or (user['roles'] == 'intern' and user['is_active'] == 0):
            session.clear()
            flash("Session expired: Your internship period has officially ended or your account has been deactivated.", "danger")
            return redirect(url_for('log_in'))

    return None

# -------------------------------------------------------------------------
# AUTHENTICATION ROUTES
# -------------------------------------------------------------------------
@app.route('/', methods=['GET', 'POST'])
def log_in():
    if 'user_id' in session:
        if session.get('role') == 'admin':
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('scan'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        if not username or not password:
            flash('Username and password are required.', 'danger')
            return render_template('login.html')
            
        with get_db_connection() as conn:
            from psycopg2.extras import RealDictCursor
            with conn.cursor(cursor_factory=RealDictCursor) as cursor:
                cursor.execute('SELECT * FROM users WHERE username = %s', (username,))
                user = cursor.fetchone()
                
        if user and user['password'] == password:
            if user['is_active'] == 0 and user['roles'] == 'intern':
                flash('Your internship period has ended.', 'danger')
                return redirect(url_for('log_in'))
                
            session.permanent = True
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['roles']
            
            if user['roles'] == 'admin':
                return redirect(url_for('admin_dashboard'))
            return redirect(url_for('scan'))
            
        flash('Invalid username or password.', 'danger')
    return render_template('login.html')



@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('log_in'))

# -------------------------------------------------------------------------
# INTERN INTERACTION ROUTES
# -------------------------------------------------------------------------
@app.route('/scan', methods=['GET', 'POST'])
def scan():
    if 'user_id' not in session:
        flash('Please log in.', 'danger')
        return redirect(url_for('log_in'))

    user_id = session['user_id']
    username = session['username']
    now = get_local_now()
    
    valid_tokens = get_current_qr_tokens()
    provided_token = request.args.get('token') or request.form.get('token')
    token_is_valid = (provided_token in valid_tokens)


    with get_db_connection() as conn:
        from psycopg2.extras import RealDictCursor
        with conn.cursor(cursor_factory=RealDictCursor) as cursor:
            # 1. Fetch last log
            cursor.execute('SELECT log_type, timestamp from logs WHERE user_id = %s ORDER BY timestamp DESC LIMIT 1', (user_id,))
            last_log = cursor.fetchone()

            # 2. Compute total credited hours
            cursor.execute('SELECT DISTINCT timestamp::date as log_date FROM logs WHERE user_id = %s', (user_id,))
            distinct_dates = cursor.fetchall()

    total_credited = sum(calculate_intern_hours(user_id, str(d['log_date']))['credited'] for d in distinct_dates)
    total_hours_formatted = round(total_credited, 2)

    # 3. Operating hours window evaluation
    sched = get_schedule_for_date(now)
    target_date_str = now.strftime('%Y-%m-%d')
    office_start = datetime.strptime(f"{target_date_str} {sched['start']}:00", '%Y-%m-%d %H:%M:%S')
    office_end = datetime.strptime(f"{target_date_str} {sched['end']}:00", '%Y-%m-%d %H:%M:%S')
    
    earliest_allowed = office_start - timedelta(hours=2)
    latest_allowed = office_end + timedelta(hours=3)
    
    outside_hours = (now < earliest_allowed or now > latest_allowed)
    allowed_window_str = f"{earliest_allowed.strftime('%I:%M %p')} - {latest_allowed.strftime('%I:%M %p')}"

    # 4. Rapid scan cooldown evaluation (30 minutes)
    cooldown_active = False
    cooldown_remaining = 0
    if last_log:
        time_delta = (now - last_log['timestamp']).total_seconds()
        if time_delta < 1800:
            cooldown_active = True
            cooldown_remaining = int((1800 - time_delta) // 60) + 1

    next_log = 'OUT' if (last_log and last_log['log_type'] == 'IN') else 'IN'

    # --- POST REQUEST: EXPLICIT BUTTON CLICK ACTION ---
    if request.method == 'POST':
        if not token_is_valid:
            flash("Invalid or expired QR code. Please scan the office monitor again.", "danger")
            return redirect(url_for('scan'))

        if outside_hours:
            flash("Action rejected: Logging is disabled outside operating hours.", "danger")
            return redirect(url_for('scan'))

        if cooldown_active:
            flash(f"Action rejected: Cooldown active. Please wait {cooldown_remaining} minute(s).", "warning")
            return redirect(url_for('scan'))


        is_late = False
        if next_log == 'IN':
            if now > office_start:
                is_late = True

        current_time_str = now.strftime('%Y-%m-%d %H:%M:%S')

        with get_db_connection() as conn:
            with conn.cursor() as cursor:
                cursor.execute('INSERT INTO logs (user_id, log_type, timestamp) VALUES (%s, %s, %s)',
                               (user_id, next_log, current_time_str))
            conn.commit()

        return render_template('confirmation.html',
                               username=username,
                               log_type=next_log,
                               time=now.strftime('%Y-%m-%d %I:%M %p'),
                               is_late=is_late,
                               outside_hours=False,
                               total_hours=total_hours_formatted)

    # --- GET REQUEST: READ-ONLY DISPLAY PORTAL ---
    last_log_time_fmt = last_log['timestamp'].strftime('%Y-%m-%d %I:%M %p') if last_log else ''

    return render_template('scan.html',
                           username=username,
                           last_log=last_log,
                           last_log_time_fmt=last_log_time_fmt,
                           next_log=next_log,
                           outside_hours=outside_hours,
                           allowed_window=allowed_window_str,
                           cooldown_active=cooldown_active,
                           cooldown_remaining=cooldown_remaining,
                           total_hours=total_hours_formatted,
                           token_is_valid=token_is_valid,
                           qr_token=provided_token)

@app.route('/display')
def display_qr():
    display_key = request.args.get('key')
    if session.get('role') != 'admin' and display_key != 'office_kiosk_123':
        return "Unauthorized", 403
    return render_template('display.html')

@app.route('/api/token')
def api_token():
    display_key = request.args.get('key')
    if session.get('role') != 'admin' and display_key != 'office_kiosk_123':
        return "Unauthorized", 403
    # Return the current most up-to-date token for the display screen
    return {"token": get_current_qr_tokens()[0]}


# -------------------------------------------------------------------------
# ADMINISTRATIVE DASHBOARD ROUTES
# -------------------------------------------------------------------------
@app.route('/admin')
def admin_dashboard():
    if session.get('role') != 'admin':
        return "Unauthorized", 403

    selected_user_id = request.args.get('filter_user', '')

    with get_db_connection() as conn:
        from psycopg2.extras import RealDictCursor
        with conn.cursor(cursor_factory=RealDictCursor) as cursor:
            # 1. Fetch Interns Roster List
            cursor.execute("SELECT id, username, is_active FROM users WHERE roles = 'intern' ORDER BY username ASC")
            interns = cursor.fetchall()

            # 2. Fetch Global Schedule System Settings
            cursor.execute("SELECT value FROM system_settings WHERE key = 'schedule_rules'")
            sched_row = cursor.fetchone()
            
            default_sched = {
                "MTW": {"start": "08:00", "end": "17:00", "break_start": "12:00", "break_end": "13:00"},
                "ThF":  {"start": "08:00", "end": "17:00", "break_start": "12:00", "break_end": "13:00"}
            }
            if sched_row:
                try:
                    sched = json.loads(sched_row['value'])
                    if not isinstance(sched, dict):
                        sched = default_sched
                    else:
                        for key in ["MTW", "ThF"]:
                            if key not in sched or not isinstance(sched[key], dict):
                                sched[key] = default_sched[key]
                            else:
                                for subkey in ["start", "end", "break_start", "break_end"]:
                                    if subkey not in sched[key]:
                                        sched[key][subkey] = default_sched[key][subkey]
                except Exception:
                    sched = default_sched
            else:
                sched = default_sched

            # 3. Build comprehensive structured timeline history matching query filters
            query = '''
                SELECT l.id, l.user_id, u.username, l.log_type, l.timestamp, l.timestamp::date as log_date 
                FROM logs l
                JOIN users u ON l.user_id = u.id
                ORDER BY l.timestamp DESC
            '''
            cursor.execute(query)
            raw_logs = cursor.fetchall()

            cursor.execute('SELECT user_id, absence_date FROM excused_absences')
            excused_list = cursor.fetchall()
            excused_map = {(e['user_id'], str(e['absence_date'])): True for e in excused_list}

            # Pre-fetch approved overtime to avoid N+1 database connection overhead
            cursor.execute('SELECT user_id, overtime_date, hours_approved FROM approved_overtime')
            ot_list = cursor.fetchall()
            ot_map = {(o['user_id'], str(o['overtime_date'])): float(o['hours_approved']) for o in ot_list}

            # Pre-fetch schedule rules to eliminate DB calls inside the calculation engine
            cursor.execute("SELECT value FROM system_settings WHERE key = 'schedule_rules'")
            sched_row = cursor.fetchone()
            if sched_row:
                sys_sched = json.loads(sched_row['value'])
            else:
                sys_sched = {}

    # Group logs by user and date in memory to pass to calculate_intern_hours
    logs_by_user_and_date = {}
    for log in raw_logs:
        uid = log['user_id']
        date_str = str(log['log_date'])
        key = (uid, date_str)
        if key not in logs_by_user_and_date:
            logs_by_user_and_date[key] = []
        logs_by_user_and_date[key].append({
            'id': log['id'],
            'log_type': log['log_type'],
            'timestamp': log['timestamp']
        })
    for key in logs_by_user_and_date:
        logs_by_user_and_date[key].reverse() # Ascending order for calculations

    # Pre-calculate full days natively to avoid O(N^2) progressive loop overhead
    day_calculations = {}
    for key, d_logs in logs_by_user_and_date.items():
        uid, date_str = key
        full_day_ot = ot_map.get((uid, date_str), 0.0)
        calc_full = calculate_intern_hours(uid, date_str, daily_logs=d_logs, ot_approved=full_day_ot, sys_sched=sys_sched)
        day_calculations[key] = calc_full

    formatted_logs = []
    user_balances = {}
    day_accumulated_credited = {}
    day_accumulated_raw = {}
    first_log_processed = {}

    # Iterate forwards chronologically (reversed raw_logs) to accurately track moving balances
    for log in reversed(raw_logs):
        uid = log['user_id']
        date_str = str(log['log_date'])
        
        calc_full = day_calculations[(uid, date_str)]
        prog = calc_full['progressive'].get(log['id'], {'raw': 0, 'credited': 0, 'potential_ot': 0})
        
        prev_cred = day_accumulated_credited.get((uid, date_str), 0.0)
        new_cred = prog['credited'] - prev_cred
        day_accumulated_credited[(uid, date_str)] = prog['credited']
        
        prev_raw = day_accumulated_raw.get((uid, date_str), 0.0)
        new_raw = prog['raw'] - prev_raw
        day_accumulated_raw[(uid, date_str)] = prog['raw']
        
        if uid not in user_balances:
            user_balances[uid] = 0.0
            
        user_balances[uid] += new_cred

        # Determine if this is the first log of the day for this user
        is_first_log = False
        if (uid, date_str) not in first_log_processed:
            first_log_processed[(uid, date_str)] = True
            is_first_log = True

        is_late = False
        if is_first_log and log['log_type'] == 'IN':
            log_sched = get_schedule_for_date(log['timestamp'], sys_sched)
            office_start = datetime.strptime(f"{date_str} {log_sched['start']}:00", '%Y-%m-%d %H:%M:%S')
            if log['timestamp'] > office_start:
                is_late = True

        formatted_logs.append({
            'id': log['id'],
            'user_id': uid,
            'username': log['username'],
            'type': log['log_type'],
            'timestamp': log['timestamp'].strftime('%Y-%m-%d %I:%M %p'),
            'date_str': date_str,
            'raw_date': log['timestamp'].strftime('%Y-%m-%d'),
            'raw_time': log['timestamp'].strftime('%H:%M'),
            'raw': round(calc_full['raw'], 2),
            'credited': round(calc_full['credited'], 2),
            'day_total_raw': round(calc_full['raw'], 2),
            'potential_ot': round(prog.get('potential_ot', 0.0), 2),
            'is_ot_approved': calc_full['is_ot_approved'],
            'approved_ot_hours': calc_full['approved_ot_hours'],
            'is_excused': excused_map.get((uid, date_str), False),
            'total_overall': round(user_balances[uid], 2),
            'is_late': is_late
        })

    # Reverse back for newest-first display on the dashboard
    formatted_logs.reverse()

    if request.args.get('partial') == '1':
        return render_template('partials/admin_content.html',
                               interns=interns,
                               logs=formatted_logs,
                               sched=sched,
                               selected_user_id=selected_user_id)

    return render_template('admin.html',
                           interns=interns,
                           logs=formatted_logs,
                           sched=sched,
                           selected_user_id=selected_user_id)

@app.route('/admin/settings', methods=['POST'])
def update_settings():
    if session.get('role') != 'admin': return "Unauthorized", 403
    
    updated_sched = {
        "MTW": {
            "start": request.form.get('mtw_start'), 
            "end": request.form.get('mtw_end'),
            "break_start": request.form.get('mtw_break_start'),
            "break_end": request.form.get('mtw_break_end')
        },
        "ThF":  {
            "start": request.form.get('thf_start'), 
            "end": request.form.get('thf_end'),
            "break_start": request.form.get('thf_break_start'),
            "break_end": request.form.get('thf_break_end')
        }
    }
    
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute('''
                INSERT INTO system_settings (key, value) VALUES ('schedule_rules', %s)
                ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value
            ''', (json.dumps(updated_sched),))
        conn.commit()
        
    flash("Schedule guidelines updated successfully.", "success")
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/intern/toggle/<int:user_id>', methods=['POST'])
def toggle_intern_status(user_id):
    if session.get('role') != 'admin': return "Unauthorized", 403
    
    status = request.form.get('is_active')
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute('UPDATE users SET is_active = %s WHERE id = %s', (int(status), user_id))
        conn.commit()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/intern/delete/<int:user_id>', methods=['POST'])
def delete_intern(user_id):
    if session.get('role') != 'admin': return "Unauthorized", 403
    
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute('DELETE FROM users WHERE id = %s', (user_id,))
        conn.commit()
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/intern/create', methods=['POST'])
def create_intern():
    if session.get('role') != 'admin': return "Unauthorized", 403
    
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    
    if not username or not password:
        flash("Username and password cannot be empty or blank spaces.", "danger")
    else:
        try:
            with get_db_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute(
                        "INSERT INTO users (username, password, roles, is_active) VALUES (%s, %s, 'intern', 1)",
                        (username, password)
                    )
                conn.commit()
            flash(f"Intern account '{username}' created successfully.", "success")
        except Exception:
            flash("Failed to create account. Username might already exist.", "danger")
            
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/logs/manual-insert', methods=['POST'])
def manual_insert_log():
    if session.get('role') != 'admin': return "Unauthorized", 403
    
    user_id = request.form.get('user_id')
    log_types = request.form.getlist('log_type[]')
    manual_dates = request.form.getlist('manual_date[]')
    manual_times = request.form.getlist('manual_time[]')
    
    # Fallback to single scalar values if [] not used in frontend form (backwards compatibility)
    if not log_types:
        log_types = [request.form.get('log_type')]
        manual_dates = [request.form.get('manual_date')]
        manual_times = [request.form.get('manual_time')]
    
    if user_id:
        with get_db_connection() as conn:
            with conn.cursor() as cursor:
                for log_type, log_date, log_time in zip(log_types, manual_dates, manual_times):
                    if log_type and log_date and log_time:
                        full_timestamp = f"{log_date} {log_time}:00"
                        cursor.execute('INSERT INTO logs (user_id, log_type, timestamp) VALUES (%s, %s, %s)',
                                       (int(user_id), log_type, full_timestamp))
            conn.commit()
            
    return redirect(url_for('admin_dashboard', filter_user=user_id))

@app.route('/admin/logs/edit', methods=['POST'])
def edit_log():
    if session.get('role') != 'admin': return "Unauthorized", 403
    
    log_id = request.form.get('log_id')
    log_type = request.form.get('log_type')
    log_date = request.form.get('manual_date')
    log_time = request.form.get('manual_time')
    user_id = request.form.get('user_id')
    
    if log_id and log_type and log_date and log_time:
        full_timestamp = f"{log_date} {log_time}:00"
        with get_db_connection() as conn:
            with conn.cursor() as cursor:
                cursor.execute('''
                    UPDATE logs 
                    SET log_type = %s, timestamp = %s 
                    WHERE id = %s
                ''', (log_type, full_timestamp, int(log_id)))
            conn.commit()
            
    return redirect(url_for('admin_dashboard', filter_user=user_id))

@app.route('/admin/logs/delete', methods=['POST'])
def delete_log():
    if session.get('role') != 'admin': return "Unauthorized", 403
    
    log_id = request.form.get('log_id')
    user_id = request.form.get('user_id')
    
    if log_id:
        with get_db_connection() as conn:
            with conn.cursor() as cursor:
                cursor.execute('DELETE FROM logs WHERE id = %s', (int(log_id),))
            conn.commit()
            
    return redirect(url_for('admin_dashboard', filter_user=user_id))

# -------------------------------------------------------------------------
# ADMINISTRATIVE OVERRIDE PATHWAYS (ABSENCE & OVERTIME)
# -------------------------------------------------------------------------
@app.route('/admin/absence/excuse', methods=['POST'])
def excuse_absence():
    if session.get('role') != 'admin': return "Unauthorized", 403
    user_id = request.form.get('user_id')
    absence_date = request.form.get('absence_date')
    
    if user_id and absence_date:
        with get_db_connection() as conn:
            with conn.cursor() as cursor:
                cursor.execute('''
                    INSERT INTO excused_absences (user_id, absence_date) VALUES (%s, %s)
                    ON CONFLICT DO NOTHING
                ''', (int(user_id), absence_date))
            conn.commit()
    return redirect(url_for('admin_dashboard', filter_user=user_id))

@app.route('/admin/absence/unexcuse', methods=['POST'])
def unexcuse_absence():
    if session.get('role') != 'admin': return "Unauthorized", 403
    user_id = request.form.get('user_id')
    absence_date = request.form.get('absence_date')
    
    if user_id and absence_date:
        with get_db_connection() as conn:
            with conn.cursor() as cursor:
                cursor.execute('DELETE FROM excused_absences WHERE user_id = %s AND absence_date = %s',
                               (int(user_id), absence_date))
            conn.commit()
    return redirect(url_for('admin_dashboard', filter_user=user_id))

@app.route('/admin/overtime/approve', methods=['POST'])
def approve_overtime():
    if session.get('role') != 'admin': return "Unauthorized", 403
    
    user_id = request.form.get('user_id')
    absence_date = request.form.get('absence_date')
    ot_hours = request.form.get('ot_hours')

    if user_id and absence_date and ot_hours:
        with get_db_connection() as conn:
            with conn.cursor() as cursor:
                cursor.execute('''
                    INSERT INTO approved_overtime (user_id, overtime_date, hours_approved)
                    VALUES (%s, %s, %s)
                    ON CONFLICT (user_id, overtime_date) DO UPDATE 
                    SET hours_approved = EXCLUDED.hours_approved
                ''', (int(user_id), absence_date, float(ot_hours)))
            conn.commit()
    return redirect(url_for('admin_dashboard', filter_user=user_id))

@app.route('/admin/overtime/revoke', methods=['POST'])
def revoke_overtime():
    if session.get('role') != 'admin': return "Unauthorized", 403
    
    user_id = request.form.get('user_id')
    absence_date = request.form.get('absence_date')

    if user_id and absence_date:
        with get_db_connection() as conn:
            with conn.cursor() as cursor:
                cursor.execute('''
                    DELETE FROM approved_overtime 
                    WHERE user_id = %s AND overtime_date = %s
                ''', (int(user_id), absence_date))
            conn.commit()
    return redirect(url_for('admin_dashboard', filter_user=user_id))

@app.route('/admin/export')
def export_excel():
    if session.get('role') != 'admin':
        return "Unauthorized", 403

    selected_user_id = request.args.get('filter_user', '')

    import csv
    from io import StringIO
    from flask import make_response

    with get_db_connection() as conn:
        from psycopg2.extras import RealDictCursor
        with conn.cursor(cursor_factory=RealDictCursor) as cursor:
            if selected_user_id:
                query = '''
                    SELECT l.id, l.user_id, u.username, l.log_type, l.timestamp, l.timestamp::date as log_date 
                    FROM logs l
                    JOIN users u ON l.user_id = u.id
                    WHERE l.user_id = %s
                    ORDER BY l.timestamp DESC
                '''
                cursor.execute(query, (selected_user_id,))
            else:
                query = '''
                    SELECT l.id, l.user_id, u.username, l.log_type, l.timestamp, l.timestamp::date as log_date 
                    FROM logs l
                    JOIN users u ON l.user_id = u.id
                    ORDER BY l.timestamp DESC
                '''
                cursor.execute(query)
                
            raw_logs = cursor.fetchall()

            cursor.execute('SELECT user_id, absence_date FROM excused_absences')
            excused_list = cursor.fetchall()
            excused_map = {(e['user_id'], str(e['absence_date'])): True for e in excused_list}

            cursor.execute('SELECT user_id, overtime_date, hours_approved FROM approved_overtime')
            ot_list = cursor.fetchall()
            ot_map = {(o['user_id'], str(o['overtime_date'])): float(o['hours_approved']) for o in ot_list}

    logs_by_user_and_date = {}
    for log in raw_logs:
        uid = log['user_id']
        date_str = str(log['log_date'])
        key = (uid, date_str)
        if key not in logs_by_user_and_date:
            logs_by_user_and_date[key] = []
        logs_by_user_and_date[key].append({
            'id': log['id'],
            'log_type': log['log_type'],
            'timestamp': log['timestamp']
        })
    for key in logs_by_user_and_date:
        logs_by_user_and_date[key].reverse()

    formatted_logs = []
    user_balances = {}
    day_accumulated_credited = {}
    day_accumulated_raw = {}
    first_log_processed = {}

    for log in reversed(raw_logs):
        uid = log['user_id']
        date_str = str(log['log_date'])
        
        daily_logs = logs_by_user_and_date.get((uid, date_str), [])
        
        current_idx = 0
        for i, dl in enumerate(daily_logs):
            if dl['id'] == log['id']:
                current_idx = i
                break
                
        progressive_logs = daily_logs[:current_idx + 1]
        is_last_log_of_day = (current_idx == len(daily_logs) - 1)
        ot_approved = ot_map.get((uid, date_str), 0.0) if is_last_log_of_day else 0.0
        
        calc = calculate_intern_hours(uid, date_str, daily_logs=progressive_logs, ot_approved=ot_approved)
        
        full_day_ot = ot_map.get((uid, date_str), 0.0)
        calc_full = calculate_intern_hours(uid, date_str, daily_logs=daily_logs, ot_approved=full_day_ot)
        
        prev_cred = day_accumulated_credited.get((uid, date_str), 0.0)
        new_cred = calc['credited'] - prev_cred
        day_accumulated_credited[(uid, date_str)] = calc['credited']
        
        prev_raw = day_accumulated_raw.get((uid, date_str), 0.0)
        new_raw = calc['raw'] - prev_raw
        day_accumulated_raw[(uid, date_str)] = calc['raw']
        
        if uid not in user_balances:
            user_balances[uid] = 0.0
            
        user_balances[uid] += new_cred

        is_first_log = False
        if (uid, date_str) not in first_log_processed:
            first_log_processed[(uid, date_str)] = True
            is_first_log = True

        is_late = False
        if is_first_log and log['log_type'] == 'IN':
            log_sched = get_schedule_for_date(log['timestamp'])
            office_start = datetime.strptime(f"{date_str} {log_sched['start']}:00", '%Y-%m-%d %H:%M:%S')
            if log['timestamp'] > office_start:
                is_late = True

        formatted_logs.append({
            'username': log['username'],
            'type': log['log_type'],
            'timestamp': log['timestamp'].strftime('%Y-%m-%d %I:%M %p'),
            'raw': round(new_raw, 2),
            'credited': round(new_cred, 2),
            'day_total_raw': round(calc_full['raw'], 2),
            'total_overall': round(user_balances[uid], 2),
            'is_late': is_late,
            'is_excused': excused_map.get((uid, date_str), False),
            'is_ot_approved': calc['is_ot_approved']
        })

    import openpyxl
    from openpyxl.styles import PatternFill
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Logs"
    ws.append(['Intern', 'Action', 'Timestamp', 'Raw Time', 'Credited Time', 'Total Balance', 'Status'])
    
    late_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    absence_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    excused_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    
    for row in reversed(formatted_logs):
        status = []
        if row['is_late']: status.append('Late')
        
        is_absence = (row['day_total_raw'] == 0 and row['type'] == 'OUT')
        if is_absence:
            if row['is_excused']:
                status.append('Excused Absence')
            else:
                status.append('Absence')
                
        if row['is_ot_approved']: status.append('OT Approved')
        status_str = ', '.join(status) if status else 'Attended'
        
        ws.append([
            row['username'],
            row['type'],
            row['timestamp'],
            row['raw'],
            row['credited'],
            row['total_overall'],
            status_str
        ])
        
        row_idx = ws.max_row
        fill_to_apply = None
        if is_absence:
            if row['is_excused']:
                fill_to_apply = excused_fill
            else:
                fill_to_apply = absence_fill
        elif row['is_late']:
            fill_to_apply = late_fill
            
        if fill_to_apply:
            for col_idx in range(1, 8):
                ws.cell(row=row_idx, column=col_idx).fill = fill_to_apply

    output = BytesIO()
    wb.save(output)
    output.seek(0)
        
    resp = make_response(output.getvalue())
    resp.headers["Content-Disposition"] = "attachment; filename=attendance_logs.xlsx"
    resp.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return resp

if __name__ == '__main__':
    # Bind to 0.0.0.0 to ensure accessibility over the local network and tunnels
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)